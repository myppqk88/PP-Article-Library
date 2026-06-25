from __future__ import annotations

import csv
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]


def _deps_for_platform() -> Path | None:
    """Pick a platform-specific deps directory if it exists.

    Layout:
      - .deps/         legacy macOS bundle (Python 3.14 darwin)
      - .deps_macos/   newer macOS bundle (preferred over .deps when present)
      - .deps_windows/ Windows-specific bundle

    On platforms without a matching dir we fall through to system site-packages.
    """
    if sys.platform == "darwin":
        for candidate in (ROOT / ".deps_macos", ROOT / ".deps"):
            if candidate.exists():
                return candidate
    elif sys.platform == "win32":
        candidate = ROOT / ".deps_windows"
        if candidate.exists():
            return candidate
    else:
        candidate = ROOT / ".deps_linux"
        if candidate.exists():
            return candidate
    return None


_deps_dir = _deps_for_platform()
if _deps_dir is not None:
    sys.path.insert(0, str(_deps_dir))

import yaml  # noqa: E402


INDEX_FIELDS = [
    "paper_id",
    "标题",
    "英文标题",
    "中文标题",
    "作者",
    "年份",
    "期刊会议",
    "期刊分区",
    "SSCI",
    "SCI",
    "UTD",
    "FT50",
    "ABS",
    "星标",
    "追踪期刊领域",
    "扫描件",
    "DOI",
    "ZoteroKey",
    "Zotero库",
    "Zotero集合",
    "Zotero标签",
    "Zotero版本",
    "AI一句话总结",
    "一级分类_AI建议",
    "二级分类_AI建议",
    "一级分类",
    "二级分类",
    "三级分类",
    "人工分类",
    "最终分类",
    "关键词",
    "研究方法",
    "研究对象",
    "与我的论文关系",
    "期刊等级_自动",
    "期刊等级_人工",
    "重要性",
    "阅读状态",
    "PDF路径",
    "笔记路径",
    "原始路径",
    "文件哈希",
    "整理时间",
    "AI模型",
    "AI置信度",
    "我的备注",
]


def load_env(path: Path | None = None) -> None:
    env_path = path or ROOT / ".env"
    if not env_path.exists():
        return
    for raw in env_path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        os.environ.setdefault(key, value)


# ------------------------------------------------------------------
# Fresh-clone bootstrap.
# After `git clone`, the user has settings.example.yaml, .env.example, and
# .gitkeep markers but NO live config and NO data dirs. Without this, the
# first call to load_settings() raises FileNotFoundError and the workbench
# fails before anyone sees the UI.
#
# This function is idempotent and cheap (~5ms when everything exists).
# Called from server.py and organize.py entry points so all CLI workflows
# get the same treatment.
# ------------------------------------------------------------------
def bootstrap_project() -> dict[str, list[str]]:
    """Ensure config + data dirs exist for first-time use.

    Returns {created: [...], copied: [...]} so callers can log what happened.
    """
    created: list[str] = []
    copied: list[str] = []

    # 1. Copy settings template if no live config
    live_settings = ROOT / "config" / "settings.yaml"
    example_settings = ROOT / "config" / "settings.example.yaml"
    if not live_settings.exists() and example_settings.exists():
        live_settings.parent.mkdir(parents=True, exist_ok=True)
        live_settings.write_text(
            example_settings.read_text(encoding="utf-8"),
            encoding="utf-8",
        )
        copied.append("config/settings.yaml")

    # 2. Copy .env template if missing
    live_env = ROOT / ".env"
    example_env = ROOT / ".env.example"
    if not live_env.exists() and example_env.exists():
        live_env.write_text(
            example_env.read_text(encoding="utf-8"),
            encoding="utf-8",
        )
        copied.append(".env")

    # 3. Create data directories.
    # These are .gitignore'd contents but the user needs them to exist.
    # mkdir(exist_ok=True) is cheap when they're already there.
    data_dirs = [
        "library/pdfs",
        "library/notes",
        "library/index",
        "library/cache",
        "library/text",
        "library/chat_history",
        "library/annotations",
        "library/stickies",
        "inbox",
        "citations",
        "exports",
        "collections/by_category",
        "prompts",
    ]
    for d in data_dirs:
        p = ROOT / d
        if not p.exists():
            p.mkdir(parents=True, exist_ok=True)
            created.append(d + "/")

    # 4. Create empty papers.csv if missing — read_csv handles missing files
    # fine but write_xlsx + render_papers_response expect at least an empty
    # CSV with header so download/sync tooling sees a real file.
    csv_path = ROOT / "library" / "index" / "papers.csv"
    if not csv_path.exists():
        header = ",".join(INDEX_FIELDS) + "\n"
        csv_path.write_text(header, encoding="utf-8-sig")
        created.append("library/index/papers.csv")

    # 5. Default prompt files if missing. Skip if user already has custom ones.
    prompts_dir = ROOT / "prompts"
    defaults = {
        "note_prompt.md": (
            "请阅读用户提供的 PDF 文本，写一份 6-10 行的中文笔记，覆盖：\n"
            "- 研究问题\n- 方法与数据\n- 主要发现\n- 局限\n- 对你正在写的论文的可借鉴之处\n"
        ),
        "classify_prompt.md": (
            "在给定的分类树里给本篇文献选一个一级分类 + 若干二级分类（可空）。\n"
            "只输出分类标签，不要解释。\n"
        ),
    }
    for fname, content in defaults.items():
        f = prompts_dir / fname
        if not f.exists():
            f.write_text(content, encoding="utf-8")
            created.append(f"prompts/{fname}")

    return {"created": created, "copied": copied}


# ------------------------------------------------------------------
# Settings cache: parse settings.yaml once, invalidate on mtime change.
# server.py called load_settings() at 36+ sites — a single /api/ask used
# to re-parse the file 3x. With cloud-synced project dirs (坚果云) the
# stat-read-parse-decode chain is non-trivial. Cache keyed on mtime so any
# external edit (or atomic_write_text) is picked up automatically.
# ------------------------------------------------------------------
_SETTINGS_CACHE: dict[str, Any] = {"mtime": None, "data": None}


def load_settings(deep: bool = True) -> dict[str, Any]:
    """Load (and cache) settings.yaml.

    By default returns a fresh deepcopy so callers can mutate freely. Hot
    loops that only READ (e.g. `easyscholar_refresh_all` iterating 506 rows)
    can pass `deep=False` to get the cached dict directly — saves a deepcopy
    per iteration. DO NOT mutate when deep=False, or you'll poison the cache
    for every other caller until the next mtime change.
    """
    settings_path = ROOT / "config" / "settings.yaml"
    try:
        mtime = settings_path.stat().st_mtime
    except OSError:
        # Fall through to a fresh parse so the caller still gets a useful error
        return yaml.safe_load(settings_path.read_text(encoding="utf-8"))
    if _SETTINGS_CACHE["mtime"] == mtime and _SETTINGS_CACHE["data"] is not None:
        if not deep:
            return _SETTINGS_CACHE["data"]  # caller promises read-only
        import copy
        return copy.deepcopy(_SETTINGS_CACHE["data"])
    data = yaml.safe_load(settings_path.read_text(encoding="utf-8"))
    _SETTINGS_CACHE["mtime"] = mtime
    _SETTINGS_CACHE["data"] = data
    if not deep:
        return data
    import copy
    return copy.deepcopy(data)


def invalidate_settings_cache() -> None:
    """Call after any direct settings.yaml write that bypasses load_settings()
    (e.g. an external editor). Atomic writes through save_settings_yaml bump
    the mtime automatically so this is rarely needed."""
    _SETTINGS_CACHE["mtime"] = None
    _SETTINGS_CACHE["data"] = None


# ------------------------------------------------------------------
# Default model names per provider.
# These were duplicated across server.py (9 sites), organize.py (4 sites),
# settings.example.yaml, and the frontend. Centralized here so renaming a
# provider's flagship model is a one-line change.
# Refresh these when providers ship new defaults; bumping a default model
# here does NOT touch users' configured settings.yaml — it only matters when
# a field is missing / empty.
# ------------------------------------------------------------------
DEFAULT_MODELS = {
    "deepseek": "deepseek-v4-pro",
    "deepseek_fast": "deepseek-chat",       # text-fast variant for note generation
    "qwen": "qwen-plus",
    "qwen_vision": "qwen-vl-plus",
    "openai_compatible": "gpt-4o-mini",
    "openai_vision": "gpt-4o-mini",
    "codex_cli": "gpt-5.4",
    "claude_cli": "claude-sonnet-4-5",
}


def default_model_for(provider: str) -> str:
    """Look up the bundled default model name for a provider key."""
    if not provider:
        return ""
    return DEFAULT_MODELS.get(str(provider).strip().lower(), "")


def rel(path: Path) -> str:
    try:
        return path.resolve().relative_to(ROOT.resolve()).as_posix()
    except ValueError:
        return str(path).replace("\\", "/")


def project_path(value: str | Path) -> Path:
    if isinstance(value, str):
        value = value.replace("\\", "/")
    path = Path(value)
    if path.is_absolute():
        return path
    return ROOT / path


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def subprocess_hidden_kwargs() -> dict[str, Any]:
    """Extra kwargs for subprocess.run/Popen that suppress the console window.

    On Windows, launching a CLI (claude.cmd / codex / strings) from a GUI-less
    server process still pops a transient cmd window for each call — annoying
    and easy to close by accident (which kills the subprocess). CREATE_NO_WINDOW
    keeps it fully background. No-op on macOS / Linux.
    """
    if sys.platform == "win32":
        # subprocess.CREATE_NO_WINDOW = 0x08000000 (literal so it imports on
        # non-Windows too, where the constant doesn't exist).
        return {"creationflags": 0x08000000}
    return {}


_CSV_CACHE: dict[str, Any] = {}  # path -> {"mtime": float, "rows": list[dict]}


def read_csv(path: Path) -> list[dict[str, str]]:
    """Parse a CSV with mtime-keyed cache.

    server.py's `load_rows()` calls this on every list/sort/filter — for the
    ~500-row papers.csv that's repeated UTF-8 decode + DictReader work on every
    request. Cache invalidates when the file's mtime changes; atomic_write_text
    bumps mtime so save_rows naturally invalidates."""
    if not path.exists():
        return []
    try:
        mtime = path.stat().st_mtime
    except OSError:
        mtime = None
    key = str(path)
    cached = _CSV_CACHE.get(key)
    if cached and mtime is not None and cached["mtime"] == mtime:
        # Return a shallow copy — callers commonly mutate rows in place
        return [dict(r) for r in cached["rows"]]
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = []
        for row in reader:
            rows.append({field: (row.get(field) or "") for field in INDEX_FIELDS})
    _CSV_CACHE[key] = {"mtime": mtime, "rows": [dict(r) for r in rows]}
    return rows


def atomic_write_text(path: Path, text: str, encoding: str = "utf-8") -> None:
    """Crash-safe text write: write to .tmp, fsync, then os.replace().

    Routes ALL note / sticky / json / csv writes through here so a process
    kill / OS sleep / antivirus lock between truncation and flush can never
    leave a 0-byte file on disk. os.replace is atomic on POSIX and on NTFS.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    try:
        with open(tmp, "w", encoding=encoding, newline="\n") as f:
            f.write(text)
            f.flush()
            try:
                os.fsync(f.fileno())
            except (OSError, AttributeError):
                pass  # fsync not supported on all FS / Windows handles
        os.replace(tmp, path)
    except Exception:
        try:
            tmp.unlink()
        except OSError:
            pass
        raise


def atomic_write_bytes(path: Path, data: bytes) -> None:
    """Binary sibling of atomic_write_text for .xlsx etc."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    try:
        with open(tmp, "wb") as f:
            f.write(data)
            f.flush()
            try:
                os.fsync(f.fileno())
            except (OSError, AttributeError):
                pass
        os.replace(tmp, path)
    except Exception:
        try:
            tmp.unlink()
        except OSError:
            pass
        raise


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    """Atomic CSV write. Builds the full content in memory then swaps the file
    in one os.replace so a crash mid-write can't leave a truncated CSV."""
    import io
    buf = io.StringIO(newline="")
    writer = csv.DictWriter(buf, fieldnames=INDEX_FIELDS)
    writer.writeheader()
    for row in rows:
        writer.writerow({field: row.get(field, "") for field in INDEX_FIELDS})
    atomic_write_text(path, buf.getvalue(), encoding="utf-8-sig")


def write_xlsx(path: Path, rows: list[dict[str, str]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "papers"
    ws.append(INDEX_FIELDS)

    for row in rows:
        ws.append([row.get(field, "") for field in INDEX_FIELDS])

    header_fill = PatternFill("solid", fgColor="E8EEF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, field in enumerate(INDEX_FIELDS, start=1):
        column = get_column_letter(col_idx)
        values = [field] + [row.get(field, "") for row in rows[:200]]
        width = min(max(len(str(v)) for v in values) + 2, 48)
        ws.column_dimensions[column].width = max(width, 10)

    link_fields = {"PDF路径", "笔记路径"}
    for row_idx, row in enumerate(rows, start=2):
        for field in link_fields:
            if field not in INDEX_FIELDS:
                continue
            col_idx = INDEX_FIELDS.index(field) + 1
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                target = project_path(str(cell.value))
                if target.exists():
                    cell.hyperlink = str(target)
                    cell.style = "Hyperlink"

    # Atomic write: serialize to BytesIO first, then swap the file in
    # one go so a crash mid-save can't leave a corrupted .xlsx.
    import io
    buf = io.BytesIO()
    wb.save(buf)
    atomic_write_bytes(path, buf.getvalue())


def clean_piece(value: str, max_len: int = 80) -> str:
    value = value.strip()
    value = re.sub(r"\s+", "_", value)
    value = re.sub(r"[\\/:*?\"<>|#%{}$!`&=+@~^,;，。、《》【】（）()]", "", value)
    value = re.sub(r"_+", "_", value).strip("._- ")
    return value[:max_len].strip("._- ") or "Unknown"


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    i = 2
    while True:
        candidate = parent / f"{stem}_{i}{suffix}"
        if not candidate.exists():
            return candidate
        i += 1


def read_text_if_exists(path: Path) -> str:
    if not path.exists():
        return ""
    return path.read_text(encoding="utf-8")


def list_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        parts = []
        for item in value:
            if isinstance(item, dict):
                parts.append("；".join(f"{k}: {v}" for k, v in item.items() if v))
            else:
                parts.append(str(item))
        return "；".join(part for part in parts if part)
    return str(value)


def text_quality(text: str) -> dict[str, Any]:
    stripped = text.strip()
    words = re.findall(r"[A-Za-z]{3,}", stripped)
    return {
        "chars": len(stripped),
        "word_count": len(words),
        "line_count": len([line for line in stripped.splitlines() if line.strip()]),
    }


def scan_status_from_text(text: str) -> str:
    quality = text_quality(text)
    stripped = text.strip()
    chars = int(quality.get("chars", 0))
    words = int(quality.get("word_count", 0))
    if chars < 200:
        return "是"
    pdf_markers = [
        "/Type",
        "/XObject",
        "/Subtype",
        "/Image",
        "/Filter",
        "/FlateDecode",
        "/Length",
        "/Width",
        "/Height",
        "stream",
        "endstream",
    ]
    marker_hits = sum(stripped.count(marker) for marker in pdf_markers)
    common_words = re.findall(
        r"\b(the|and|of|in|to|for|with|that|this|is|are|as|by|from|review|research|study|paper|data|method|results?)\b",
        stripped,
        flags=re.I,
    )
    if marker_hits >= 12 and len(common_words) < 30:
        return "是"
    if marker_hits >= 6 and len(common_words) < 15:
        return "疑似"
    if chars < 1200 and words < 80:
        return "疑似"
    return ""


def normalize_journal_name(value: str) -> str:
    text = str(value or "").lower()
    text = text.replace("&", " and ")
    text = re.sub(r"\b(the|journal|of|for|and|in|on|a|an|to|with|from|by)\b", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def journal_acronym(value: str) -> str:
    stopwords = {"the", "of", "for", "and", "in", "on", "a", "an", "to", "with", "from", "by"}
    words = re.findall(r"[A-Za-z0-9]+", str(value or ""))
    kept = [word for word in words if word.lower() not in stopwords]
    return "".join(word[0].lower() for word in kept if word)


def journal_match_keys(name: str, include_acronym: bool = True) -> set[str]:
    clean = normalize_journal_name(name)
    keys = {clean}
    if include_acronym:
        keys.add(journal_acronym(name))
    compact = re.sub(r"\b(of|for|and|the)\b", " ", clean)
    compact = re.sub(r"\s+", " ", compact).strip()
    if "association information science technology" in compact:
        keys.add("jasist")
    if "american society information science technology" in compact:
        keys.add("jasist")
    return {key for key in keys if key}


def load_tracking_journals(settings: dict[str, Any]) -> list[dict[str, str]]:
    tracking = settings.get("tracking_journals", {}) or {}
    path = project_path(tracking.get("path", "list.xlsx"))
    if not path.exists():
        return []
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_name = tracking.get("sheet", "") or wb.sheetnames[0]
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(min_row=1, max_row=20, max_col=8, values_only=True)
    header_row = 1
    headers: list[str] = []
    for row_no, row in enumerate(rows, start=1):
        values = [str(cell).strip() if cell is not None else "" for cell in row]
        if "期刊名称" in values:
            header_row = row_no
            headers = values
            break
    if not headers:
        headers = [str(cell).strip() if cell is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1, max_col=8, values_only=True))]

    def col_index(*names: str) -> int | None:
        for name in names:
            if name in headers:
                return headers.index(name)
        return None

    name_idx = col_index("期刊名称", "Journal", "Journal Name", "journal")
    field_idx = col_index("领域口径", "领域", "学科")
    ft50_idx = col_index("FT50")
    utd_idx = col_index("UTD24", "UTD")
    abs_idx = col_index("ABS3/4标记", "ABS", "ABS标记")
    abs_subject_idx = col_index("ABS学科", "ABS Subject", "ABS Subject Area")
    if name_idx is None:
        return []

    entries: list[dict[str, str]] = []
    blank_streak = 0
    for row in ws.iter_rows(min_row=header_row + 1, max_col=max(len(headers), 6), values_only=True):
        values = ["" if cell is None else str(cell).strip() for cell in row]
        if not any(values[:6]):
            blank_streak += 1
            if blank_streak >= 200:
                break
            continue
        blank_streak = 0
        name = values[name_idx] if name_idx < len(values) else ""
        if not name:
            continue
        entries.append(
            {
                "name": name,
                "field": values[field_idx] if field_idx is not None and field_idx < len(values) else "",
                "ft50": values[ft50_idx] if ft50_idx is not None and ft50_idx < len(values) else "",
                "utd": values[utd_idx] if utd_idx is not None and utd_idx < len(values) else "",
                "abs": values[abs_idx] if abs_idx is not None and abs_idx < len(values) else "",
                "abs_subject": values[abs_subject_idx] if abs_subject_idx is not None and abs_subject_idx < len(values) else "",
            }
        )
    return entries


def save_tracking_journals(settings: dict[str, Any], entries: list[dict[str, str]]) -> Path:
    tracking = settings.get("tracking_journals", {}) or {}
    path = project_path(tracking.get("path", "list.xlsx"))
    sheet_name = tracking.get("sheet", "SSCI_Q1_扩展初筛") or "SSCI_Q1_扩展初筛"
    path.parent.mkdir(parents=True, exist_ok=True)

    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill

    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
    if sheet_name in wb.sheetnames:
        index = wb.sheetnames.index(sheet_name)
        old = wb[sheet_name]
        wb.remove(old)
        ws = wb.create_sheet(sheet_name, index)
    else:
        if len(wb.sheetnames) == 1 and wb.active.max_row == 1 and wb.active.max_column == 1 and wb.active["A1"].value is None:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name, 0)

    headers = ["领域口径", "期刊名称", "FT50", "UTD24", "ABS3/4标记", "ABS学科"]
    ws.append(headers)
    for entry in entries:
        name = str(entry.get("name", "")).strip()
        if not name:
            continue
        ws.append(
            [
                str(entry.get("field", "")).strip(),
                name,
                str(entry.get("ft50", "")).strip(),
                str(entry.get("utd", "")).strip(),
                str(entry.get("abs", "")).strip(),
                str(entry.get("abs_subject", "")).strip(),
            ]
        )

    header_fill = PatternFill("solid", fgColor="E8EEF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    widths = [28, 48, 10, 10, 14, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    wb.save(path)
    return path


def match_tracking_journal(venue: str, entries: list[dict[str, str]]) -> dict[str, str] | None:
    if not venue or not entries:
        return None
    candidate_parts = [venue]
    candidate_parts.extend(re.split(r"[;,；，|/:：]", venue))
    candidates: set[str] = set()
    for part in candidate_parts:
        words = re.findall(r"[A-Za-z0-9]+", part)
        compact = re.sub(r"[^A-Za-z0-9]", "", part)
        include_acronym = bool(compact) and len(compact) <= 12 and (part.strip().isupper() or len(words) <= 2)
        candidates.update(journal_match_keys(part, include_acronym=include_acronym))
    candidates = {item for item in candidates if len(item) >= 3}
    if not candidates:
        return None

    key_map: dict[str, dict[str, str]] = {}
    for entry in entries:
        for key in journal_match_keys(entry.get("name", "")):
            if len(key) >= 3:
                key_map.setdefault(key, entry)
    for candidate in candidates:
        if candidate in key_map:
            return key_map[candidate]

    normalized_venue = normalize_journal_name(venue)
    for key, entry in sorted(key_map.items(), key=lambda item: len(item[0]), reverse=True):
        if len(key) >= 10 and (normalized_venue.startswith(key) or key.startswith(normalized_venue)):
            return entry
    return None


def yes_like(value: str) -> bool:
    return str(value or "").strip() in {"是", "yes", "Yes", "YES", "Y", "y", "1", "true", "True"}


def apply_tracking_journal(row: dict[str, str], entries: list[dict[str, str]]) -> bool:
    before = {field: row.get(field, "") for field in ["星标", "追踪期刊领域", "FT50", "UTD", "ABS"]}
    match = match_tracking_journal(row.get("期刊会议", ""), entries)
    if match:
        row["星标"] = "★"
        row["追踪期刊领域"] = match.get("field", "")
        if yes_like(match.get("ft50", "")):
            row["FT50"] = "是"
        if yes_like(match.get("utd", "")):
            row["UTD"] = "是"
        abs_value = str(match.get("abs", "")).strip()
        if abs_value in {"是", "需核验"}:
            row["ABS"] = abs_value
    else:
        row["星标"] = ""
        row["追踪期刊领域"] = ""
    after = {field: row.get(field, "") for field in ["星标", "追踪期刊领域", "FT50", "UTD", "ABS"]}
    return before != after


def apply_tracking_journals(rows: list[dict[str, str]], settings: dict[str, Any]) -> dict[str, int]:
    entries = load_tracking_journals(settings)
    changed = 0
    starred = 0
    for row in rows:
        if apply_tracking_journal(row, entries):
            changed += 1
        if row.get("星标"):
            starred += 1
    return {"changed": changed, "starred": starred, "entries": len(entries)}


def sort_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    def key(row: dict[str, str]) -> tuple[str, str, str]:
        return (row.get("年份", ""), row.get("作者", ""), row.get("标题", ""))

    return sorted(rows, key=key, reverse=True)


# Display-side sort used by /api/papers. The CSV / xlsx on disk are still
# written in the legacy 年份-desc order via `sort_rows` above, so the human-
# readable spreadsheet keeps its old layout; the web list can be reordered
# independently.

# Order阅读状态 by reading-progress: newer/AI-only first, "已读" / "暂不相关" last.
_READ_STATUS_ORDER = {
    "AI初整": 0,
    "待读": 1,
    "已读": 2,
    "精读": 3,
    "可引用": 4,
    "暂不相关": 5,
    "": 9,  # 未定 sinks to the bottom in asc, top in desc
}


def _row_sort_key(row: dict[str, str], sort_by: str) -> tuple:
    if sort_by == "added":
        # 整理时间 is "YYYY-MM-DD HH:MM:SS" string; lexicographic order works.
        return (row.get("整理时间", ""), row.get("paper_id", ""))
    if sort_by == "year":
        return (row.get("年份", ""), row.get("作者", ""), row.get("标题", ""))
    if sort_by == "importance":
        raw = (row.get("重要性", "") or "").strip()
        try:
            value = int(raw)
        except ValueError:
            value = -1  # 未定 always goes last
        return (value, row.get("整理时间", ""))
    if sort_by == "read_status":
        status = (row.get("阅读状态", "") or "").strip()
        return (_READ_STATUS_ORDER.get(status, 8), row.get("整理时间", ""))
    # default fallback: 添加时间
    return (row.get("整理时间", ""), row.get("paper_id", ""))


SUPPORTED_VIEW_SORTS = ("added", "year", "importance", "read_status")


def sort_rows_for_view(
    rows: list[dict[str, str]],
    sort_by: str = "added",
    order: str = "desc",
) -> list[dict[str, str]]:
    """Return a new list of rows sorted for display in the web UI.

    sort_by ∈ {"added", "year", "importance", "read_status"}.
    order ∈ {"asc", "desc"}; default "desc" (most useful in every case:
    newest-added on top, highest year on top, highest importance on top,
    earliest reading-state on top — i.e. AI初整/待读 before 已读).
    """
    key = sort_by if sort_by in SUPPORTED_VIEW_SORTS else "added"
    reverse = (order or "desc").lower() != "asc"
    return sorted(rows, key=lambda row: _row_sort_key(row, key), reverse=reverse)
